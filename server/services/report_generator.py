#!/usr/bin/env python3
"""
Security report generator for PDF and Excel formats
"""

import json
import sys
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
import io

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Preformatted
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: ReportLab not available, PDF generation will be limited")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available, Excel generation will be limited")

class SecurityReportGenerator:
    def __init__(self):
        self.styles = None
        if REPORTLAB_AVAILABLE:
            self._setup_styles()
    
    def _setup_styles(self):
        """Setup custom styles for PDF generation"""
        self.styles = getSampleStyleSheet()
        
        # Custom styles
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            textColor=colors.HexColor('#1f2937'),
            alignment=1  # Center
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#374151'),
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomBody',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6,
        ))
        
        self.styles.add(ParagraphStyle(
            name='CodeStyle',
            parent=self.styles['Normal'],
            fontSize=9,
            fontName='Courier',
            leftIndent=20,
            backgroundColor=colors.HexColor('#f3f4f6'),
        ))

    def _get_vuln_field(self, vuln: Dict[str, Any], *keys, default: Optional[Any] = None) -> Any:
        """Return the first present key from vuln among provided keys.

        This helps accept different key styles (camelCase / snake_case).
        """
        for k in keys:
            if isinstance(vuln, dict) and k in vuln and vuln[k] is not None:
                return vuln[k]
        return default
    
    def generate_report(self, scan_data: Dict[str, Any], report_config: Dict[str, Any]) -> str:
        """Generate security report in specified format"""
        
        report_format = report_config.get('format', 'pdf').lower()
        output_path = f"reports/security_report_{scan_data['scan']['id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if report_format == 'pdf':
            return self._generate_pdf_report(scan_data, report_config, f"{output_path}.pdf")
        elif report_format == 'excel':
            return self._generate_excel_report(scan_data, report_config, f"{output_path}.xlsx")
        else:
            raise ValueError(f"Unsupported report format: {report_format}")
    
    def _generate_pdf_report(self, scan_data: Dict[str, Any], config: Dict[str, Any], output_path: str) -> str:
        """Generate PDF security report"""
        
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError("ReportLab is required for PDF generation. Please install it using: pip install reportlab")
        
        # Ensure reports directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        story = []
        
        # Title
        title = Paragraph("Security Vulnerability Report", self.styles['CustomTitle'])
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Executive Summary
        if 'executive_summary' in config.get('sections', []):
            story.extend(self._create_executive_summary(scan_data))
        
        # Vulnerability Summary
        story.extend(self._create_vulnerability_summary(scan_data))
        
        # Detailed Findings
        if 'vulnerability_details' in config.get('sections', []):
            story.extend(self._create_detailed_findings(scan_data, config))
        
        # AI Recommendations
        if 'ai_recommendations' in config.get('sections', []):
            story.extend(self._create_ai_recommendations(scan_data))
        
        # Build PDF
        doc.build(story)
        return output_path
    
    def _generate_excel_report(self, scan_data: Dict[str, Any], config: Dict[str, Any], output_path: str) -> str:
        """Generate Excel security report"""
        
        if not OPENPYXL_AVAILABLE:
            return self._generate_csv_report(scan_data, config, output_path.replace('.xlsx', '.csv'))
        
        # Ensure reports directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet and create custom sheets
        wb.remove(wb.active)
        
        # Summary sheet
        self._create_excel_summary_sheet(wb, scan_data)
        
        # Vulnerabilities sheet
        self._create_excel_vulnerabilities_sheet(wb, scan_data, config)
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def _create_executive_summary(self, scan_data: Dict[str, Any]) -> List:
        """Create executive summary section for PDF"""
        story = []
        
        story.append(Paragraph("Executive Summary", self.styles['CustomHeading']))
        
        scan = scan_data['scan']
        vulnerabilities = scan_data['vulnerabilities']
        
        # Summary statistics
        total_vulns = len(vulnerabilities)
        severity_counts = self._count_by_severity(vulnerabilities)
        
        summary_text = f"""
        <b>Scan Overview:</b><br/>
        • Scan ID: {scan['id']}<br/>
        • Date: {scan['createdAt']}<br/>
        • Files Scanned: {scan['totalFiles']}<br/>
        • Total Vulnerabilities: {total_vulns}<br/><br/>
        
        <b>Severity Breakdown:</b><br/>
        • Critical: {severity_counts.get('critical', 0)}<br/>
        • High: {severity_counts.get('high', 0)}<br/>
        • Medium: {severity_counts.get('medium', 0)}<br/>
        • Low: {severity_counts.get('low', 0)}<br/>
        """
        
        story.append(Paragraph(summary_text, self.styles['CustomBody']))
        story.append(Spacer(1, 20))
        
        return story
    
    def _create_vulnerability_summary(self, scan_data: Dict[str, Any]) -> List:
        """Create vulnerability summary table for PDF"""
        story = []
        
        story.append(Paragraph("Vulnerability Summary", self.styles['CustomHeading']))
        
        vulnerabilities = scan_data['vulnerabilities']
        
        if not vulnerabilities:
            story.append(Paragraph("No vulnerabilities detected.", self.styles['CustomBody']))
            return story
        
        # Create table data
        table_data = [['File', 'Line', 'Type', 'Severity', 'CWE']]
        
        for vuln in vulnerabilities[:20]:  # Limit to top 20 for summary
            filename = self._get_vuln_field(vuln, 'filename', 'file', 'fileName')
            line = str(self._get_vuln_field(vuln, 'lineNumber', 'line_number', 'line') or '')
            vtype = self._get_vuln_field(vuln, 'vulnerabilityType', 'vulnerability_type', 'type')
            severity = self._get_vuln_field(vuln, 'severity', 'severity')
            cwe = self._get_vuln_field(vuln, 'cweId', 'cwe_id', 'cwe') or 'N/A'

            table_data.append([
                filename,
                line,
                vtype,
                severity.title() if isinstance(severity, str) else str(severity),
                cwe
            ])
        
        # Create and style table
        table = Table(table_data, colWidths=[2*inch, 0.7*inch, 2*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
        
        return story
    
    def _create_detailed_findings(self, scan_data: Dict[str, Any], config: Dict[str, Any]) -> List:
        """Create detailed findings section for PDF"""
        story = []
        
        story.append(PageBreak())
        story.append(Paragraph("Detailed Vulnerability Findings", self.styles['CustomHeading']))
        
        vulnerabilities = scan_data['vulnerabilities']
        severity_filter = config.get('severityFilter', 'all')
        
        # Filter vulnerabilities based on severity
        filtered_vulns = self._filter_vulnerabilities_by_severity(vulnerabilities, severity_filter)
        
        for i, vuln in enumerate(filtered_vulns[:10], 1):  # Limit to top 10 detailed
            vtype = self._get_vuln_field(vuln, 'vulnerabilityType', 'vulnerability_type', 'type')
            filename = self._get_vuln_field(vuln, 'filename', 'file', 'fileName')
            line = self._get_vuln_field(vuln, 'lineNumber', 'line_number', 'line')
            severity = self._get_vuln_field(vuln, 'severity', 'severity')
            cwe = self._get_vuln_field(vuln, 'cweId', 'cwe_id', 'cwe') or 'N/A'
            tool = self._get_vuln_field(vuln, 'detectionTool', 'detection_tool', 'tool')
            confidence = self._get_vuln_field(vuln, 'confidence', 'confidence')
            description = self._get_vuln_field(vuln, 'description', 'desc', '')

            story.append(Paragraph(f"Finding {i}: {vtype}", self.styles['CustomHeading']))

            details = f"""
            <b>File:</b> {filename}<br/>
            <b>Line:</b> {line}<br/>
            <b>Severity:</b> {str(severity).title()}<br/>
            <b>CWE ID:</b> {cwe}<br/>
            <b>Detection Tool:</b> {tool}<br/>
            <b>Confidence:</b> {confidence}%<br/><br/>

            <b>Description:</b><br/>
            {description}
            """

            story.append(Paragraph(details, self.styles['CustomBody']))

            if config.get('includeCode', True):
                # Get code snippet from multiple possible keys
                code_snippet = self._get_vuln_field(vuln, 'codeSnippet', 'code_snippet', 'code', '')
                story.append(Paragraph("<b>Code Snippet:</b>", self.styles['CustomBody']))
                if code_snippet:
                    # Use Preformatted if available to preserve indentation
                    try:
                        story.append(Preformatted(code_snippet, self.styles['CodeStyle']))
                    except Exception:
                        # Fallback to Paragraph with the CodeStyle
                        safe = code_snippet.replace('<', '&lt;').replace('>', '&gt;')
                        story.append(Paragraph(safe, self.styles['CodeStyle']))
                else:
                    story.append(Paragraph("No code snippet available.", self.styles['CodeStyle']))

            story.append(Spacer(1, 15))
        
        return story
    
    def _create_ai_recommendations(self, scan_data: Dict[str, Any]) -> List:
        """Create AI recommendations section for PDF"""
        story = []
        
        story.append(PageBreak())
        story.append(Paragraph("AI-Generated Recommendations", self.styles['CustomHeading']))
        
        recommendations = [
            "Implement parameterized queries for all database operations to prevent SQL injection",
            "Use environment variables for storing sensitive configuration data",
            "Implement proper input validation and output encoding",
            "Use strong cryptographic functions (bcrypt, scrypt, or Argon2) for password hashing",
            "Regular security code reviews and automated scanning",
            "Keep dependencies updated and monitor for known vulnerabilities"
        ]
        
        rec_text = "<br/>".join([f"• {rec}" for rec in recommendations])
        story.append(Paragraph(rec_text, self.styles['CustomBody']))
        
        return story
    
    def _create_excel_summary_sheet(self, workbook, scan_data: Dict[str, Any]):
        """Create Excel summary sheet"""
        ws = workbook.create_sheet("Summary", 0)
        
        # Headers
        ws['A1'] = "Security Scan Summary"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Scan information
        scan = scan_data['scan']
        ws['A3'] = "Scan ID:"
        ws['B3'] = scan['id']
        ws['A4'] = "Date:"
        ws['B4'] = scan['createdAt']
        ws['A5'] = "Files Scanned:"
        ws['B5'] = scan['totalFiles']
        
        # Severity summary
        vulnerabilities = scan_data['vulnerabilities']
        severity_counts = self._count_by_severity(vulnerabilities)
        
        ws['A7'] = "Vulnerability Breakdown"
        ws['A7'].font = Font(size=14, bold=True)
        
        ws['A8'] = "Critical:"
        ws['B8'] = severity_counts.get('critical', 0)
        ws['A9'] = "High:"
        ws['B9'] = severity_counts.get('high', 0)
        ws['A10'] = "Medium:"
        ws['B10'] = severity_counts.get('medium', 0)
        ws['A11'] = "Low:"
        ws['B11'] = severity_counts.get('low', 0)
        
        # Styling
        for row in range(8, 12):
            severity = ws[f'A{row}'].value.replace(':', '').lower()
            if severity == 'critical':
                ws[f'B{row}'].fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
            elif severity == 'high':
                ws[f'B{row}'].fill = PatternFill(start_color="FFF2E5", end_color="FFF2E5", fill_type="solid")
    
    def _create_excel_vulnerabilities_sheet(self, workbook, scan_data: Dict[str, Any], config: Dict[str, Any]):
        """Create Excel vulnerabilities sheet"""
        ws = workbook.create_sheet("Vulnerabilities", 1)
        
        # Headers
        headers = ['File', 'Line', 'Type', 'Severity', 'CWE', 'Description', 'Code Snippet', 'Tool', 'Confidence']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        
        # Data
        vulnerabilities = scan_data['vulnerabilities']
        severity_filter = config.get('severityFilter', 'all')
        filtered_vulns = self._filter_vulnerabilities_by_severity(vulnerabilities, severity_filter)
        
        for row, vuln in enumerate(filtered_vulns, 2):
            ws.cell(row=row, column=1, value=vuln['filename'])
            ws.cell(row=row, column=2, value=vuln['lineNumber'])
            ws.cell(row=row, column=3, value=vuln['vulnerabilityType'])
            ws.cell(row=row, column=4, value=vuln['severity'].title())
            ws.cell(row=row, column=5, value=vuln.get('cweId', 'N/A'))
            # Include code snippet (support multiple key names)
            code_snip = self._get_vuln_field(vuln, 'codeSnippet', 'code_snippet', 'code', '')
            ws.cell(row=row, column=6, value=vuln['description'])
            ws.cell(row=row, column=7, value=code_snip)
            ws.cell(row=row, column=8, value=vuln['detectionTool'])
            ws.cell(row=row, column=9, value=f"{vuln['confidence']}%")
            
            # Color code by severity
            severity_colors = {
                'critical': 'FFE5E5',
                'high': 'FFF2E5',
                'medium': 'FFFBF2',
                'low': 'F0FDF4'
            }
            color = severity_colors.get(vuln['severity'], 'FFFFFF')
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _count_by_severity(self, vulnerabilities: List[Dict[str, Any]]) -> Dict[str, int]:
        """Count vulnerabilities by severity"""
        counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
        for vuln in vulnerabilities:
            severity = vuln.get('severity', 'low').lower()
            if severity in counts:
                counts[severity] += 1
        return counts
    
    def _filter_vulnerabilities_by_severity(self, vulnerabilities: List[Dict[str, Any]], severity_filter: str) -> List[Dict[str, Any]]:
        """Filter vulnerabilities based on severity"""
        if severity_filter == 'all':
            return vulnerabilities
        elif severity_filter == 'critical_high':
            return [v for v in vulnerabilities if v['severity'].lower() in ['critical', 'high']]
        elif severity_filter == 'medium_above':
            return [v for v in vulnerabilities if v['severity'].lower() in ['critical', 'high', 'medium']]
        else:
            return vulnerabilities
    
    def _generate_text_report(self, scan_data: Dict[str, Any], config: Dict[str, Any], output_path: str) -> str:
        """Generate simple text report as fallback"""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        with open(output_path, 'w') as f:
            f.write("SECURITY VULNERABILITY REPORT\n")
            f.write("=" * 50 + "\n\n")
            
            scan = scan_data['scan']
            f.write(f"Scan ID: {scan['id']}\n")
            f.write(f"Date: {scan['createdAt']}\n")
            f.write(f"Files Scanned: {scan['totalFiles']}\n\n")
            
            vulnerabilities = scan_data['vulnerabilities']
            severity_counts = self._count_by_severity(vulnerabilities)
            
            f.write("VULNERABILITY SUMMARY:\n")
            f.write("-" * 25 + "\n")
            f.write(f"Critical: {severity_counts['critical']}\n")
            f.write(f"High: {severity_counts['high']}\n")
            f.write(f"Medium: {severity_counts['medium']}\n")
            f.write(f"Low: {severity_counts['low']}\n\n")
            
            if vulnerabilities:
                f.write("DETAILED FINDINGS:\n")
                f.write("-" * 20 + "\n")
                for i, vuln in enumerate(vulnerabilities, 1):
                    f.write(f"{i}. {vuln['vulnerabilityType']} ({vuln['severity'].upper()})\n")
                    f.write(f"   File: {vuln['filename']}:{vuln['lineNumber']}\n")
                    f.write(f"   CWE: {vuln.get('cweId', 'N/A')}\n")
                    f.write(f"   Description: {vuln['description']}\n\n")
        
        return output_path
    
    def _generate_csv_report(self, scan_data: Dict[str, Any], config: Dict[str, Any], output_path: str) -> str:
        """Generate CSV report as fallback for Excel"""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        import csv
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            writer.writerow(['File', 'Line', 'Type', 'Severity', 'CWE', 'Description', 'Tool', 'Confidence'])
            
            # Data
            vulnerabilities = scan_data['vulnerabilities']
            for vuln in vulnerabilities:
                writer.writerow([
                    vuln['filename'],
                    vuln['lineNumber'],
                    vuln['vulnerabilityType'],
                    vuln['severity'],
                    vuln.get('cweId', 'N/A'),
                    vuln['description'],
                    vuln['detectionTool'],
                    f"{vuln['confidence']}%"
                ])
        
        return output_path

def main():
    if len(sys.argv) < 3:
        print("Usage: python report_generator.py <scan_data_json> <config_json>")
        sys.exit(1)
    
    try:
        scan_data = json.loads(sys.argv[1])
        config = json.loads(sys.argv[2])
        
        generator = SecurityReportGenerator()
        report_path = generator.generate_report(scan_data, config)
        
        print(json.dumps({
            'success': True,
            'report_path': report_path,
            'message': f'Report generated successfully at {report_path}'
        }))
        
    except Exception as e:
        print(json.dumps({
            'success': False,
            'error': str(e)
        }))
        sys.exit(1)

if __name__ == "__main__":
    main()
